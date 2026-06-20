#!/usr/bin/env python3
"""
aggregate_judges.py — Combine scores from 5 human judges into a single expert score.

For each image:
  1. Collect the 5 individual scores.
  2. Drop the highest and the lowest.
  3. Average the remaining 3.
  4. Round UP to one decimal place (ceiling at 1 d.p.).

Reads:
  data/subset_puntuation_juez{1..5}.xlsx   — POSITION, IMAGE, SCORE

Writes:
  data/subset_puntuation_expert.xlsx       — POSITION, IMAGE, SCORE

Usage:
    python aggregate_judges.py [--data-dir PATH] [--output PATH]
"""
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent

N_JUDGES = 5
JUDGE_FILES = [
    ROOT / "data" / f"subset_puntuation_juez{i}.xlsx" for i in range(1, N_JUDGES + 1)
]


def _extract_image_name(cell_value: str | None) -> str | None:
    if cell_value is None:
        return None
    s = str(cell_value)
    m = re.search(r'[/\\]([^/\\"]+\.(?:jpg|jpeg|png|JPG|JPEG|PNG))', s)
    return m.group(1) if m else s.strip()


def load_scores(path: Path) -> dict[tuple[str, str], float]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    scores: dict[tuple[str, str], float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pos, img, score = row[0], row[1], row[2]
        if pos is None or img is None or score is None:
            continue
        img_name = _extract_image_name(str(img))
        if img_name:
            key = (str(pos).strip().lower(), img_name.upper())
            scores[key] = float(score)
    return scores


def _ceil_1dec(x: float) -> float:
    return math.ceil(round(x, 9) * 10) / 10


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--data-dir", type=Path, default=ROOT / "data",
        help="Directory containing the judge Excel files",
    )
    parser.add_argument(
        "--output", type=Path,
        default=ROOT / "data" / "subset_puntuation_expert.xlsx",
        help="Output Excel file path",
    )
    args = parser.parse_args()

    judge_files = [args.data_dir / f"subset_puntuation_juez{i}.xlsx" for i in range(1, N_JUDGES + 1)]

    # Load all judge score maps
    all_maps: list[dict[tuple[str, str], float]] = []
    for path in judge_files:
        if not path.exists():
            print(f"ERROR: judge file not found: {path}")
            return 1
        all_maps.append(load_scores(path))
        print(f"Loaded {len(all_maps[-1])} entries from {path.name}")

    # Use juez1 as the reference for row order and original display values
    ref_wb = openpyxl.load_workbook(judge_files[0], data_only=False)
    ref_ws = ref_wb.active

    ref_rows: list[tuple[str, str]] = []  # (position_display, image_display)
    for row in ref_ws.iter_rows(min_row=2, values_only=True):
        pos, img = row[0], row[1]
        if pos is None or img is None:
            continue
        img_display = _extract_image_name(str(img)) or str(img).strip()
        ref_rows.append((str(pos).strip(), img_display))

    # Aggregate
    results: list[tuple[str, str, float]] = []
    skipped = 0
    for pos_display, img_display in ref_rows:
        key = (pos_display.lower(), img_display.upper())
        scores_for_key: list[float] = []
        for judge_map in all_maps:
            if key in judge_map:
                scores_for_key.append(judge_map[key])

        if len(scores_for_key) < 3:
            print(f"SKIP — fewer than 3 scores for {img_display} ({pos_display}), got {len(scores_for_key)}")
            skipped += 1
            continue

        scores_sorted = sorted(scores_for_key)
        if len(scores_sorted) == 5:
            middle = scores_sorted[1:4]
        elif len(scores_sorted) == 4:
            middle = scores_sorted[1:3]
        else:
            middle = scores_sorted

        mean_score = sum(middle) / len(middle)
        final_score = _ceil_1dec(mean_score)
        results.append((pos_display, img_display, final_score))

    print(f"\n{len(results)} images processed, {skipped} skipped.")

    # Write output Excel
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Scores"

    header_font = Font(bold=True)
    out_ws.append(["POSITION", "IMAGE", "SCORE"])
    for cell in out_ws[1]:
        cell.font = header_font

    for pos, img, score in results:
        out_ws.append([pos, img, score])

    # Column widths
    out_ws.column_dimensions["A"].width = 35
    out_ws.column_dimensions["B"].width = 20
    out_ws.column_dimensions["C"].width = 10

    args.output.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(args.output)
    print(f"Saved -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
